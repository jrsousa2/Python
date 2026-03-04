import openpyxl
from os.path import exists, isfile

# OPENS AN EXCEL FILE AND SPECIFIED SHEET
# LAYOUT ARE THE COLS OF THE FILE
def open_excel(Arq,Sheet, data_only_pmt=True):
    dict = {}
    dict["file"] = None
    dict["sheet"] = None
    headers = None
    if not isfile(Arq):
       print("File doesn't exist") 
    else:
        try:
            workbook = openpyxl.load_workbook(Arq, data_only=data_only_pmt)
            worksheet = workbook[Sheet]
        except:
            # IF AN EXCEPTION OCCURS, ALL WILL BE NONE
            print("Sheet doesn't exist!") 
            dict["file"] = None  
            dict["sheet"] = None 
        else:
            dict["file"] = workbook
            worksheet = workbook.active
            dict["sheet"] = worksheet
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
    # HEADERS WILL BE NONE
    dict["headers"] = headers    
    return dict

# FINDS THE COL NUMBER OF A COLUMN
def col_number(headers,col_name):
    if col_name in headers:
       col = headers.index(col_name)+1
    else:
        col = ""   
    return col 

# FINDS LAST ROW OF THE SHEET
def last_row(sheet):
   last_row = None
   if sheet is not None:
        # Iterate through the rows in column A to find the last non-empty cell
        last_row = None

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            cell = row[0]  # Column A (index 0)
            if cell.value is not None:
                last_row = cell.row
   else:
       print("Worksheet does not exist!")             
   return last_row

# PLACEHOLDER?
def read_from_excel(worksheet,nrow,col_number):
    nrow = 1
    return nrow 

# WRITES TO OPEN EXCEL FILE
def write_to_excel(worksheet,nrow,headers,col_name,value_to_write):
    col = headers.index(col_name)
    worksheet.cell(row=nrow, column=col, value=value_to_write)
    nrow = nrow+1
    return nrow

# FIRST EMPTY ROW IN THE FILE
def empty_row(worksheet):
    next_row = 2
    while worksheet.cell(row=next_row, column=1).value is not None:
          next_row = next_row+1
    
    return next_row  

# Function to search for a value in a column and 
# return another column value from the same row
def search_df(df, search_col, search_val, return_col):
    result = df.loc[df[search_col] == search_val, return_col]
    if not result.empty:
        return result.iloc[0]
    else:
        return None