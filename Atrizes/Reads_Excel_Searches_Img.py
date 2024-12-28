# READS AN EXCEL FILE WITH ACTOR NAMES
# LOOKS FOR THE PATHS OF IMAGES NAMED AFTER THE ACTORS IN A DIR
# IGNORES NUMBERS AND SPACES AT THE END OF THE IMAGE NAME.

#from requests import get
from os.path import exists, isfile
import openpyxl
from unidecode import unidecode
from re import sub

import sys 
  
# Insert the path of modules folder  
sys.path.insert(0, "D:\\iTunes\\Codes") 
# import Read_PL # type: ignore

import Files # type: ignore

# VARIABLE USED IN THE iTunes FUNCTIONS
# ExtArray = [".unk",".jpg",".png",".bmp"]


# OPENS AN EXCEL FILE AND SPECIFIED SHEET
# LAYOUT ARE THE COLS OF THE FILE
def open_excel(Arq,Sheet):
    dict = {}
    dict["file"] = None
    dict["sheet"] = None
    if not isfile(Arq):
       print("File doesn't exist") 
    else:
        workbook = openpyxl.load_workbook(Arq) #, data_only=True
        worksheet = workbook[Sheet]
        worksheet = workbook.active
    dict["file"] = workbook
    dict["sheet"] = worksheet   
    headers = []
    for cell in worksheet[1]:
        headers.append(cell.value)
    dict["headers"] = headers
    return dict

# FIRST EMPTY ROW IN THE FILE
def empty_row(worksheet):
    next_row = 2
    while worksheet.cell(row=next_row, column=1).value is not None:
          next_row = next_row+1
    
    return next_row      

def last_row(sheet):
   # Iterate through the rows in column A to find the last non-empty cell
   last_row = None

   for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
      cell = row[0]  # Column A (index 0)
      if cell.value is not None:
         last_row = cell.row
   return last_row 

# WRITE TO ARTIST EXCEL Art_layout = ["API", "Search", "Cur_Artist", "Nbr_srch_art", "Srch_Artists", "Srch_Variations"]
def col_number(headers,col_name):
    if col_name in headers:
       col = headers.index(col_name)+1
    else:
        col = ""   
    return col 

# WRITE TO ARTIST EXCEL Art_layout = ["API", "Search", "Cur_Artist", "Nbr_srch_art", "Srch_Artists", "Srch_Variations"]
def read_from_excel(worksheet,nrow,col_number):
    nrow = 1
    return nrow  

# Function to search for a value in a column and return another column value from the same row
def search_df(df, search_col, search_val, return_col):
    result = df.loc[df[search_col] == search_val, return_col]
    if not result.empty:
        return result.iloc[0]
    else:
        return None

# Normalize a string by removing spaces and trailing numbers
def normalize(s):
    return sub(r'\s|\d+$', '', s)

# MAIN CODE
def Pop_images():

    # OPENS EXCEL FILE TO READ FROM (WRITES TO THE SAME FILE)
    Excel_infile = r"D:\Videos\Atrizes BR\Atrizes.xlsx"
    Excel_outfile = r"D:\Videos\Atrizes BR\Atrizes_out.xlsx"
    excelf = open_excel(Excel_infile,"Atrizes")
    workbook = excelf["file"]
    worksheet = excelf["sheet"]
    headers = excelf["headers"]
    rows = last_row(worksheet)-1

    # Artist	Title	Type
    Atriz_col = col_number(headers,"Atriz")
    Sel_col = col_number(headers,"Select")
    Img_col = col_number(headers,"Image")
    #Novela_col = col_number(headers,"Novela")
    #Year_col = col_number(headers,"Year")

    # BUILDS IMAGE LIST
    # ext = ['.jpg', '.png', '.gif']
    # LIST OF IMAGES
    list = Files.get_Win_files(r"D:\Videos\Atrizes BR\Pics", ['.jpg', '.png'])

    # LOOP        
    next_row = 1
    while next_row+1 <= rows+1:
          next_row = next_row+1
          Atriz_value = worksheet.cell(row=next_row, column=Atriz_col).value
          Sel_value = worksheet.cell(row=next_row, column=Sel_col).value
          Img_value = worksheet.cell(row=next_row, column=Img_col).value

          if Sel_value =="x" and (Img_value is None or not exists(Img_value)):
             srch_key = unidecode(Atriz_value.lower())
             
             # Search for "atriz" in y and return x
             Image_path = [x for x, y in list if normalize(srch_key) in normalize(y)]
             if len(Image_path)>0:
                print("Atriz",Atriz_value,"Image:",Image_path[0],"(",len(Image_path),"images)\n")
                worksheet.cell(row=next_row, column=Img_col, value=Image_path[0])
                # workbook.save(Excel_infile)
             
    # SAVE SHEET
    workbook.save(Excel_infile)

# CALLS FUNC
Pop_images()
 
                   