# READS AN EXCEL FILE WITH URL'S FOR WIKIPEDIA PAGES
# PARSES THE LOADED HTML PAGES AND CAPTURES TEXT

#from requests import get
from os.path import exists, isfile
import openpyxl
from unidecode import unidecode
import requests
from bs4 import BeautifulSoup

import sys
sys.path.insert(0, "D:\\Python\\Modules")

import Read_PL
import Excel

# VARIABLE USED IN THE iTunes FUNCTIONS
ExtArray = [".unk",".jpg",".png",".bmp"]

def Read_URL(url):
    # Fetch the page content
    response = requests.get(url)

    if response.status_code == 200:  # Success
        # Process the content
        page_text = response.text

        # Parse the content with BeautifulSoup
        soup = BeautifulSoup(response.content, "html.parser")

        # Find the table or section containing the names and characters
        table = soup.find('table', {'class': 'wikitable sortable'})

        if table is None:
           table = soup.find('table', {'class': 'wikitable'}) 

        
        # Extracting rows of the table
        actor_list = []
        if table is not None:
           th_tags = table.find_all('th')
           #Pers_pos = th_tags.index("<th>Personagem</th>")
           Pers_pos = next((i for i, th in enumerate(th_tags) if th.get_text(strip=True) == "Personagem"), -1)

           for row in table.find_all('tr'):
               columns = row.find_all('td')
               cols = len(columns)
               if cols>=2:  # Assuming two columns (Ator, Personagem)
                  try:
                     actor = columns[0].get_text(strip=True)
                     if actor == "":
                        actor = columns[1].get_text(strip=True) 
                     if actor == "":
                        actor = columns[2].get_text(strip=True)
                     if Pers_pos>0:   
                        character = columns[Pers_pos].get_text(strip=True)
                     else:
                         character = columns[cols-1].get_text(strip=True)   
                  except:
                     # REST
                     for col in columns:
                         text = col.get_text(strip=True)
                         if text:  # Check if the cell is not empty
                            actor = text
                            break  # Exit after finding the first non-empty cell
                     character = "N/A"
                    
                  # THE END
                  dict = {"Actor": actor, "Character": character}
                  print(f'{actor}: {character}')
                  actor_list.append(dict)
        print()        
    else:
        print(f"Failed to retrieve page. Status code: {response.status_code}")
    
    return actor_list

# MAIN CODE
def Load_pages(PL_name=None,PL_nbr=None):

    # XLS FILE TO WRITE TO
    wExcel_file = "D:\\Videos\\Atrizes BR\\Results2.xlsx"
    wexcelf = Excel.open_excel(wExcel_file,"Sheet1")
    wworksheet = wexcelf["sheet"]
    wheaders = wexcelf["headers"]
    wnext_row = Excel.last_row(wworksheet)

    # Artist	Title	Type
    wNovPos_col = Excel.col_number(wheaders,"#")
    wCount_col = Excel.col_number(wheaders,"Count")
    wActor_col = Excel.col_number(wheaders,"Atriz")
    wChar_col = Excel.col_number(wheaders,"Personagem")
    wNovela_col = Excel.col_number(wheaders,"Novela")
    wYear_col = Excel.col_number(wheaders,"Ano")

    # OPENS EXCEL FILE TO READ FROM
    Excel_file = "D:\\Videos\\Atrizes BR\\Brasil.xlsm"
    excelf = Excel.open_excel(Excel_file,"FALTA")
    worksheet = excelf["sheet"]
    headers = excelf["headers"]
    rows = Excel.last_row(worksheet)-1

    # Artist	Title	Type
    URL_col = Excel.col_number(headers,"URL")
    NovPos_col = Excel.col_number(headers,"#")
    Novela_col = Excel.col_number(headers,"Novela")
    Year_col = Excel.col_number(headers,"Year")

    # WHERE TO START?
    last_count = wworksheet.cell(row=wnext_row, column=wCount_col).value
    if type(last_count)==int:
       next_row = last_count+1
    else:
        next_row = 1

    # LOOP        
    nbr = 0
    while next_row+1 <= rows+1:
          next_row = next_row+1
          URL_value = worksheet.cell(row=next_row, column=URL_col).value
          NovPos_value = worksheet.cell(row=next_row, column=NovPos_col).value
          Novela_value = worksheet.cell(row=next_row, column=Novela_col).value
          Year_value = worksheet.cell(row=next_row, column=Year_col).value

          # READ
          actor_list = Read_URL(URL_value)

          # LIST
          nbr_actors = len(actor_list)
          print("Novela:",Novela_value,"// Actors:",nbr_actors,"\n")

          # WRITES TO SHEET
          for i in range(nbr_actors):
              actor = actor_list[i]["Actor"]
              character = actor_list[i]["Character"]
              wnext_row = wnext_row+1
              wworksheet.cell(row=wnext_row, column=wNovPos_col, value=NovPos_value)
              wworksheet.cell(row=wnext_row, column=wCount_col, value=next_row-1)
              wworksheet.cell(row=wnext_row, column=wActor_col, value=actor)
              wworksheet.cell(row=wnext_row, column=wChar_col, value=character)
              wworksheet.cell(row=wnext_row, column=wNovela_col, value=Novela_value)
              wworksheet.cell(row=wnext_row, column=wYear_col, value=Year_value)

          # SAVE SHEET
          if wnext_row % 1==0:
             wexcelf["file"].save(wExcel_file)

# CALLS FUNC
Load_pages()
 
                   