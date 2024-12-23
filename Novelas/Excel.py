import openpyxl
#from requests import get
from os.path import exists, isfile

# OPENS AN EXCEL FILE AND SPECIFIED SHEET
# LAYOUT ARE THE COLS OF THE FILE
def open_excel(Arq,Sheet):
    dict = {}
    dict["file"] = None
    dict["sheet"] = None
    if not isfile(Arq):
       print("File doesn't exist") 
    else:
        workbook = openpyxl.load_workbook(Arq) #(Arq, data_only=True
        worksheet = workbook[Sheet]
        # worksheet = workbook.active
    dict["file"] = workbook
    dict["sheet"] = worksheet   
    headers = []
    for cell in worksheet[1]:
        headers.append(cell.value)
    dict["headers"] = headers
    return dict

# FIRST EMPTY ROW IN THE FILE
def empty_row(worksheet):
    next_row = 2
    while worksheet.cell(row=next_row, column=1).value is not None:
          next_row = next_row+1
    
    return next_row      

def last_row(sheet):
   # Iterate through the rows in column A to find the last non-empty cell
   last_row = None

   for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
      cell = row[0]  # Column A (index 0)
      if cell.value is not None:
         last_row = cell.row
   return last_row 

# WRITE TO ARTIST EXCEL Art_layout = ["API", "Search", "Cur_Artist", "Nbr_srch_art", "Srch_Artists", "Srch_Variations"]
def col_number(headers,col_name):
    if col_name in headers:
       col = headers.index(col_name)+1
    else:
        col = ""   
    return col 

# WRITE TO ARTIST EXCEL Art_layout = ["API", "Search", "Cur_Artist", "Nbr_srch_art", "Srch_Artists", "Srch_Variations"]
def read_from_excel(worksheet,nrow,col_number):
    nrow = 1
    return nrow  

# Function to search for a value in a column and return another column value from the same row
def search_df(df, search_col, search_val, return_col):
    result = df.loc[df[search_col] == search_val, return_col]
    if not result.empty:
        return result.iloc[0]
    else:
        return None