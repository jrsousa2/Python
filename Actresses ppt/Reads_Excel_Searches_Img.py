# READS AN EXCEL FILE WITH ACTOR NAMES
# LOOKS FOR THE PATHS OF IMAGES NAMED AFTER THE ACTORS IN A DIRECTORY WITH SAVED IMAGES
# IGNORES NUMBERS AND SPACES AT THE END OF THE IMAGE NAME
# CODE DETECTED ALL IMAGES EXCEPT FOR A FEW TYPOS IN THE ACTOR NAMES
# AFTER THIS CODE IS RUN, THE EXCEL FILE IS USED AS INPUT IN THE PPT CODE THAT ADDS SLIDES

#from requests import get
from os.path import exists, isfile
import openpyxl
from unidecode import unidecode
from re import sub

import sys
sys.path.insert(0, "D:\\Python\\Modules")

import Excel
import Files

# VARIABLE USED IN THE iTunes FUNCTIONS
# ExtArray = [".unk",".jpg",".png",".bmp"]


# Function to search for a value in a column and return another column value from the same row
def search_df(df, search_col, search_val, return_col):
    result = df.loc[df[search_col] == search_val, return_col]
    if not result.empty:
        return result.iloc[0]
    else:
        return None

# Normalize a string by removing spaces and trailing numbers
def normalize(s):
    return sub(r'\s|\d+$', '', s)

# MAIN CODE
def Pop_images():

    # OPENS EXCEL FILE TO READ FROM (WRITES TO THE SAME FILE)
    Excel_infile = r"D:\Videos\Atrizes BR\Atrizes.xlsx"
    Excel_outfile = r"D:\Videos\Atrizes BR\Atrizes_out.xlsx"
    excelf = Excel.open_excel(Excel_infile,"Atrizes")
    workbook = excelf["file"]
    worksheet = excelf["sheet"]
    headers = excelf["headers"]
    rows = Excel.last_row(worksheet)-1

    # Artist	Title	Type
    Atriz_col = Excel.col_number(headers,"Atriz")
    Sel_col = Excel.col_number(headers,"Select")
    Img_col = Excel.col_number(headers,"Image")
    #Novela_col = col_number(headers,"Novela")
    #Year_col = col_number(headers,"Year")

    # BUILDS IMAGE LIST
    # ext = ['.jpg', '.png', '.gif']
    # LIST OF IMAGES
    list = Files.get_Win_files(r"D:\Videos\Atrizes BR\Pics", ['.jpg', '.png'])

    # LOOP        
    next_row = 1
    while next_row+1 <= rows+1:
          next_row = next_row+1
          Atriz_value = worksheet.cell(row=next_row, column=Atriz_col).value
          Sel_value = worksheet.cell(row=next_row, column=Sel_col).value
          Img_value = worksheet.cell(row=next_row, column=Img_col).value

          if Sel_value =="x" and (Img_value is None or not exists(Img_value)):
             srch_key = unidecode(Atriz_value.lower())
             
             # Search for "atriz" in y and return x
             Image_path = [x for x, y in list if normalize(srch_key) in normalize(y)]
             if len(Image_path)>0:
                print("Atriz",Atriz_value,"Image:",Image_path[0],"(",len(Image_path),"images)\n")
                worksheet.cell(row=next_row, column=Img_col, value=Image_path[0])
                # workbook.save(Excel_infile)
             
    # SAVE SHEET
    workbook.save(Excel_infile)

# CALLS FUNC
Pop_images()
 
                   