# READS AN EXCEL FILE AND SEARCHES FOR THE MP3 FILES REFERENCED IN IT BY THE TAGS
# SAVES THE FULL PATH OF THE FOUND FILE IN THE SPREADSHEET

#from requests import get
from os.path import exists, isfile
import openpyxl
import os
import fnmatch
#import re
#from unidecode import unidecode

import sys
sys.path.insert(0, "D:\\Python\\Modules")

import Excel
import Read_PL
import Files
from Search import similar_ratio
#from Stdz import Simple_stdz

# SPLITS THE FILE LIST INTO A DICTIONARY
def Create_dict(list):
    result_dict = {}

    for item in list:
        if item.find(" - ")>=0 and item.count(" - ")==1:
           # Split the item into key and value using the "-" separator
           key, value = item.split(" - ")
           
           # If the key is not already in the dictionary, create a new entry
           key = key.lower()
           if key not in result_dict:
              result_dict[key] = []
           
           # Append the value to the corresponding key in the dictionary
           result_dict[key].append(item.replace(".mp3",""))

    # Print the resulting dictionary
    return result_dict

# FINDS A FILE IN ROOT DIRECTORY 
# RETURNS TRUE OR FALSE
def file_in_dir(root_dir, target_file):
    for foldername, subfolders, filenames in os.walk(root_dir):
        if target_file in filenames:
           # return os.path.join(foldername, target_file)
           return True
    return False

# FINDS FULL FILENAME target in root_dir
def find_fullfile(root_dir, target):
    full_file = ""
    if target != "":
       matching_files = [os.path.join(root, file) for root, dir, files in os.walk(root_dir) for file in fnmatch.filter(files, target)]
       if len(matching_files)>0:
          full_file = matching_files[0]
    # RETURNS FILE FOUND
    return full_file

# MAIN CODE
def Updt_novela(PL_name=None,PL_nbr=None):

    dict = Read_PL.Init_iTunes()
    playlists = dict['PLs'] 

    root_dir = "E:\\MP3"
   
    # OPENS EXCEL FILE
    #Excel_file = "D:\\Python\\Excel\\Novela_selected.xlsx"
    Excel_file = "D:\\Videos\\Novelas80s.xlsx"
    excelf = Excel.open_excel(Excel_file,"Final")

    worksheet = excelf["sheet"]
    headers = excelf["headers"]

    rows = Excel.last_row(worksheet)-1

    # LIST OF ALL FILE NAMES WO/ THE PATH
    print("Creating list of all files...\n")
    all_filenames = []
    for foldername, subfolders, filenames in os.walk(root_dir):
        all_filenames.extend(filenames)
    all_filenames = [x.lower() for x in all_filenames]  

    # LIST OF ALL FILE NAMES WO/ THE PATH THAT HAVE NOVELA IN THE GENRE
    print("Creating list of mp3 files wo/ extension...\n")
    all_files_mp3 = []
    no_files = len(all_filenames)
    cnt = 0
    for filename in all_filenames:
        cnt = cnt+1
        if cnt % 1000 == 0:
           print("Checking file",cnt,"of",no_files)
        if filename.lower().find(".mp3")>=0:
           all_files_mp3.append(Files.file_wo_ext(filename))
    # all_files_mp3 = [x.lower() for x in all_files_mp3] 
    print()

    # file_dict = Create_dict(all_files_mp3)

    # CRIA PLAYLIST APENAS SE HOUVER ARQUIVOS PRA ATUALIZAR
    PL_novela_nm = "Updt_novela"
    PL_novela = Read_PL.Create_PL(PL_novela_nm,recreate="n") 

    # Artist	Title	Type
    File_col = Excel.col_number(headers,"File")
    Art_col = Excel.col_number(headers,"Artist")
    Title_col = Excel.col_number(headers,"Title")
    Genre_col = Excel.col_number(headers,"Genre")
    Type_col = Excel.col_number(headers,"Type")
    Path_col = Excel.col_number(headers,"Fullname")
    Score_col = Excel.col_number(headers,"Score")
    # col1 = col_number(headers,"Key1")
    next_row = 1
    found = 0
    searched = 0
    #worksheet.cell(row=next_row+1, column=col1).value is not None and worksheet.cell(row=next_row+1, column=Type_col).value is None
    while next_row+1 <= rows:
          next_row = next_row+1
          Path_value = worksheet.cell(row=next_row, column=Path_col).value
          File_value = worksheet.cell(row=next_row, column=File_col).value
          delim_count = File_value.count(" - ")
          if Path_value is None and delim_count>=1:
             searched = searched+1
             File_txt = worksheet.cell(row = next_row, column=File_col).value
             File_txt = File_txt.replace(".mp3","")
             File_lst = File_txt.split(" - ")
             Title = File_lst[delim_count-1].strip()
             Art = File_lst[delim_count].strip()
             print("Checking",next_row-1,"of",rows,"tracks")
             filename = Art + " - " + Title + ".mp3"
             if filename.lower() in all_filenames:
                print("File found:",filename)
                worksheet.cell(row=next_row, column=Type_col, value="Art/Title")
                worksheet.cell(row=next_row, column=Art_col, value=Art) 
                worksheet.cell(row=next_row, column=Title_col, value=Title)
             else:
                filename = Title + " - " + Art + ".mp3"
                if filename.lower() in all_filenames:
                   print("File found:",filename)
                   worksheet.cell(row=next_row, column=Type_col, value="Title/Art")
                   worksheet.cell(row=next_row, column=Art_col, value=Title) 
                   worksheet.cell(row=next_row, column=Title_col, value=Art)
                else:
                    filename = Art + " " + Title
                    print("File not found:",filename)
                    # words = filename.split()
                    base_len = len(filename.replace(" ",""))
                    # common_words(filename,base_len,'richard sanderson, vladimir cosma - reality')
                    # SCAN LIST AND STORE NUMBER OF COMMON WORDS
                    dicts_lst = [similar_ratio(filename,x) for x in all_files_mp3]
                    ratios_lst = [x['ratio'] for x in dicts_lst]
                    ratio_obs = max(ratios_lst)
                    pos = ratios_lst.index(ratio_obs)
                    if ratio_obs > 0.6:
                       # BEST MATCH
                       best_lst = all_files_mp3[pos].split(" - ")
                       Art = best_lst[0].title()
                       Title = best_lst[1].title()
                       print("Best match:",all_files_mp3[pos].title(),"// score:", ratio_obs)
                       worksheet.cell(row=next_row, column=Type_col, value="Fuzzy")
                       worksheet.cell(row=next_row, column=Art_col, value=Art) 
                       worksheet.cell(row=next_row, column=Title_col, value=Title)
                       worksheet.cell(row=next_row, column=Score_col, value=ratio_obs)
                       filename = all_files_mp3[pos] + ".mp3"
                       # excelf["file"].save(Excel_file)
             # LOOKS FOR THE FILE
             full_filename = find_fullfile(root_dir, filename)
             if exists(full_filename):
                found = found + 1
                print("Path:",full_filename)
                print("Total",searched,") Found:",found,"\\ Not found:",searched-found)
                worksheet.cell(row=next_row, column=Path_col, value=full_filename)
                if found % 20==0:
                   excelf["file"].save(Excel_file)
                Read_PL.Add_file_to_PL(playlists,PL_novela_nm,full_filename)
             print()
          elif worksheet.cell(row=next_row, column=Genre_col).value is None:
               full_filename = worksheet.cell(row=next_row, column=Path_col).value
               try:
                  Genre = Files.read_eyed3(full_filename,"Genre").name # MP3_tag
               except:
                   pass
               else:   
                   worksheet.cell(row=next_row, column=Genre_col, value=Genre)
               if next_row % 40==0:
                  excelf["file"].save(Excel_file)
               # excelf["file"].save(Excel_file)

    # SAVE MISMATCH EXCEL FILES AGAIN
    print("Saving excel file...")
    excelf["file"].save(Excel_file)

   
# CALLS FUNC
Updt_novela()
 
                   